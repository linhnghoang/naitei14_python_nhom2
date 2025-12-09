from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Q, Count
from datetime import datetime
import re

from ..models import Category, Book, Publisher


DEFAULT_CATEGORY_COLUMNS = [
    "id",
    "name",
    "slug",
    "description",
    "parent_name",
    "books_count",
    "children_count",
    "hierarchy_level",
]


DEFAULT_PUBLISHER_COLUMNS = [
    "id",
    "name",
    "description",
    "founded_year",
    "website",
    "books_count",
    "created_at",
]


def build_category_queryset(params, include_books=False):
    """Build category queryset based on filter parameters."""
    qs = Category.objects.all()

    # Search by name or description
    if "q" in params and params["q"].strip():
        q = params["q"].strip()
        qs = qs.filter(
            Q(name__icontains=q) | Q(description__icontains=q) | Q(slug__icontains=q)
        )

    # Filter by parent category
    if "parent_id" in params and params["parent_id"]:
        try:
            parent_id = int(params["parent_id"])
            if parent_id == 0:  # Top-level categories
                qs = qs.filter(parent=None)
            else:
                qs = qs.filter(parent_id=parent_id)
        except (ValueError, TypeError):
            pass

    # Filter by minimum books count
    if "min_books" in params and params["min_books"]:
        try:
            min_books = int(params["min_books"])
            qs = qs.annotate(books_count=Count("books")).filter(
                books_count__gte=min_books
            )
        except (ValueError, TypeError):
            pass

    # Filter empty categories
    if "empty_only" in params and params["empty_only"].lower() in ["true", "1", "yes"]:
        qs = qs.filter(books=None)

    # Filter categories with subcategories
    if "has_children" in params and params["has_children"].lower() in [
        "true",
        "1",
        "yes",
    ]:
        qs = qs.filter(children__isnull=False).distinct()
    elif "has_children" in params and params["has_children"].lower() in [
        "false",
        "0",
        "no",
    ]:
        qs = qs.filter(children__isnull=True)

    # Sorting
    sort = params.get("sort", "name")
    sort_mapping = {
        "name": "name",
        "-name": "-name",
        "books_count": "books_count",  # Will need annotation
        "-books_count": "-books_count",
        "children_count": "children_count",  # Will need annotation
        "-children_count": "-children_count",
        "id": "id",
        "-id": "-id",
    }

    if sort in sort_mapping:
        order_field = sort_mapping[sort]

        # Add annotations for count-based sorting
        if "books_count" in order_field:
            qs = qs.annotate(books_count=Count("books", distinct=True))
        if "children_count" in order_field:
            qs = qs.annotate(children_count=Count("children", distinct=True))

        qs = qs.order_by(order_field)

    # Optimize with select_related and prefetch_related
    qs = qs.select_related("parent")
    if include_books:
        qs = qs.prefetch_related("books", "children")
    else:
        qs = qs.prefetch_related("children")

    return qs


def calculate_hierarchy_level(category):
    """Calculate the hierarchy level of a category (0 = top level)."""
    level = 0
    parent = category.parent
    while parent:
        level += 1
        parent = parent.parent
    return level


def get_category_hierarchy_path(category):
    """Get the full hierarchy path of a category."""
    path = [category.name]
    parent = category.parent
    while parent:
        path.insert(0, parent.name)
        parent = parent.parent
    return " > ".join(path)


def render_categories_workbook(queryset, columns=None, include_books=False):
    """Create an Excel workbook with categories data."""
    columns = columns or DEFAULT_CATEGORY_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    column_headers = {
        "id": "ID",
        "name": "Name",
        "slug": "Slug",
        "description": "Description",
        "parent_name": "Parent Category",
        "books_count": "Books Count",
        "children_count": "Subcategories Count",
        "hierarchy_level": "Level",
        "hierarchy_path": "Full Path",
    }

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_headers.get(col_name, col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    row_idx = 2
    for category in queryset:
        # Annotate with counts if not already done
        books_count = getattr(category, "books_count", category.books.count())
        children_count = getattr(category, "children_count", category.children.count())

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name == "id":
                cell.value = category.id
            elif col_name == "name":
                cell.value = category.name
            elif col_name == "slug":
                cell.value = category.slug
            elif col_name == "description":
                cell.value = category.description or ""
            elif col_name == "parent_name":
                cell.value = category.parent.name if category.parent else ""
            elif col_name == "books_count":
                cell.value = books_count
            elif col_name == "children_count":
                cell.value = children_count
            elif col_name == "hierarchy_level":
                cell.value = calculate_hierarchy_level(category)
            elif col_name == "hierarchy_path":
                cell.value = get_category_hierarchy_path(category)

        row_idx += 1

    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add books sheet if requested
    if include_books:
        books_ws = wb.create_sheet("Books by Category")

        # Books sheet headers
        books_headers = [
            "Category ID",
            "Category Name",
            "Category Path",
            "Book ID",
            "Book Title",
            "ISBN13",
            "Publisher",
            "Year",
        ]
        for col_idx, header in enumerate(books_headers, 1):
            cell = books_ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write books data
        books_row = 2
        for category in queryset:
            category_path = get_category_hierarchy_path(category)
            books = category.books.all().select_related("publisher")

            if books:
                for book in books:
                    books_ws.cell(row=books_row, column=1).value = category.id
                    books_ws.cell(row=books_row, column=2).value = category.name
                    books_ws.cell(row=books_row, column=3).value = category_path
                    books_ws.cell(row=books_row, column=4).value = book.id
                    books_ws.cell(row=books_row, column=5).value = book.title
                    books_ws.cell(row=books_row, column=6).value = book.isbn13 or ""
                    books_ws.cell(row=books_row, column=7).value = (
                        book.publisher.name if book.publisher else ""
                    )
                    books_ws.cell(row=books_row, column=8).value = (
                        book.publish_year or ""
                    )
                    books_row += 1
            else:
                # Add category row even if no books
                books_ws.cell(row=books_row, column=1).value = category.id
                books_ws.cell(row=books_row, column=2).value = category.name
                books_ws.cell(row=books_row, column=3).value = category_path
                books_ws.cell(row=books_row, column=4).value = "No books"
                books_row += 1

        # Auto-adjust books sheet column widths
        for col_idx in range(1, len(books_headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in books_ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            books_ws.column_dimensions[column_letter].width = adjusted_width

    return wb


def build_book_queryset(params, include_items=False):
    """Build book queryset based on filter parameters - for compatibility
    with base project."""
    qs = Book.objects.all()

    # Search by title, description, or ISBN
    if "q" in params and params["q"].strip():
        q = params["q"].strip()
        qs = qs.filter(
            Q(title__icontains=q)
            | Q(description__icontains=q)
            | Q(isbn13__icontains=q)
        )

    # Filter by category
    if "category_id" in params and params["category_id"]:
        try:
            category_id = int(params["category_id"])
            qs = qs.filter(categories__id=category_id)
        except (ValueError, TypeError):
            pass

    # Filter by author
    if "author_id" in params and params["author_id"]:
        try:
            author_id = int(params["author_id"])
            qs = qs.filter(authors__id=author_id)
        except (ValueError, TypeError):
            pass

    # Filter by publisher
    if "publisher_id" in params and params["publisher_id"]:
        try:
            publisher_id = int(params["publisher_id"])
            qs = qs.filter(publisher_id=publisher_id)
        except (ValueError, TypeError):
            pass

    # Filter by language
    if "language" in params and params["language"].strip():
        qs = qs.filter(language_code__icontains=params["language"].strip())

    # Date range filters
    if "created_from" in params and params["created_from"]:
        try:
            from_date = datetime.strptime(params["created_from"], "%Y-%m-%d").date()
            qs = qs.filter(created_at__date__gte=from_date)
        except (ValueError, TypeError):
            pass

    if "created_to" in params and params["created_to"]:
        try:
            to_date = datetime.strptime(params["created_to"], "%Y-%m-%d").date()
            qs = qs.filter(created_at__date__lte=to_date)
        except (ValueError, TypeError):
            pass

    # Optimize with select_related and prefetch_related
    qs = qs.select_related("publisher").prefetch_related("authors", "categories")
    if include_items:
        qs = qs.prefetch_related("items")

    return qs.distinct()


def render_books_workbook(queryset, columns=None, include_items=False):
    """Create an Excel workbook with books data - for compatibility
    with base project."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # Basic implementation for books export
    headers = ["ID", "Title", "ISBN13", "Publisher", "Year", "Categories", "Authors"]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = header

    # Write data
    for row_idx, book in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1).value = book.id
        ws.cell(row=row_idx, column=2).value = book.title
        ws.cell(row=row_idx, column=3).value = book.isbn13 or ""
        ws.cell(row=row_idx, column=4).value = (
            book.publisher.name if book.publisher else ""
        )
        ws.cell(row=row_idx, column=5).value = book.publish_year or ""
        ws.cell(row=row_idx, column=6).value = ", ".join(
            [cat.name for cat in book.categories.all()]
        )
        ws.cell(row=row_idx, column=7).value = ", ".join(
            [auth.name for auth in book.authors.all()]
        )

    return wb


def build_publisher_queryset(params, include_books=False):
    """Build publisher queryset based on filter parameters."""
    qs = Publisher.objects.all()

    # Search by name, description, or website
    if "q" in params and params["q"].strip():
        q = params["q"].strip()
        qs = qs.filter(
            Q(name__icontains=q) | Q(description__icontains=q) | Q(website__icontains=q)
        )

    # Filter by founded year range
    if "founded_year_from" in params and params["founded_year_from"]:
        try:
            year_from = int(params["founded_year_from"])
            qs = qs.filter(founded_year__gte=year_from)
        except (ValueError, TypeError):
            pass

    if "founded_year_to" in params and params["founded_year_to"]:
        try:
            year_to = int(params["founded_year_to"])
            qs = qs.filter(founded_year__lte=year_to)
        except (ValueError, TypeError):
            pass

    # Filter by minimum books count
    if "min_books" in params and params["min_books"]:
        try:
            min_books = int(params["min_books"])
            qs = qs.annotate(books_count=Count("books")).filter(
                books_count__gte=min_books
            )
        except (ValueError, TypeError):
            pass

    # Filter publishers without books
    if "empty_only" in params and params["empty_only"].lower() in ["true", "1", "yes"]:
        qs = qs.filter(books__isnull=True)

    # Filter by website presence
    if "has_website" in params and params["has_website"].lower() in [
        "true",
        "1",
        "yes",
    ]:
        qs = qs.exclude(Q(website="") | Q(website__isnull=True))
    elif "has_website" in params and params["has_website"].lower() in [
        "false",
        "0",
        "no",
    ]:
        qs = qs.filter(Q(website="") | Q(website__isnull=True))

    # Date range filters for creation date
    if "created_from" in params and params["created_from"]:
        try:
            from_date = datetime.strptime(params["created_from"], "%Y-%m-%d").date()
            qs = qs.filter(created_at__date__gte=from_date)
        except (ValueError, TypeError):
            pass

    if "created_to" in params and params["created_to"]:
        try:
            to_date = datetime.strptime(params["created_to"], "%Y-%m-%d").date()
            qs = qs.filter(created_at__date__lte=to_date)
        except (ValueError, TypeError):
            pass

    # Sorting
    sort = params.get("sort", "name")
    sort_mapping = {
        "name": "name",
        "-name": "-name",
        "founded_year": "founded_year",
        "-founded_year": "-founded_year",
        "books_count": "books_count",  # Will need annotation
        "-books_count": "-books_count",
        "created_at": "created_at",
        "-created_at": "-created_at",
        "id": "id",
        "-id": "-id",
    }

    if sort in sort_mapping:
        order_field = sort_mapping[sort]

        # Add annotations for count-based sorting
        if "books_count" in order_field:
            qs = qs.annotate(books_count=Count("books", distinct=True))

        qs = qs.order_by(order_field)

    # Optimize with prefetch_related
    if include_books:
        qs = qs.prefetch_related("books")

    return qs


def render_publishers_workbook(queryset, columns=None, include_books=False):
    """Create an Excel workbook with publishers data."""
    columns = columns or DEFAULT_PUBLISHER_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Publishers"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    column_headers = {
        "id": "ID",
        "name": "Name",
        "description": "Description",
        "founded_year": "Founded Year",
        "website": "Website",
        "books_count": "Books Count",
        "created_at": "Created At",
        "years_active": "Years Active",
    }

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_headers.get(col_name, col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    row_idx = 2
    current_year = datetime.now().year

    for publisher in queryset:
        # Annotate with counts if not already done
        books_count = getattr(publisher, "books_count", publisher.books.count())

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name == "id":
                cell.value = publisher.id
            elif col_name == "name":
                cell.value = publisher.name
            elif col_name == "description":
                cell.value = publisher.description or ""
            elif col_name == "founded_year":
                cell.value = publisher.founded_year or ""
            elif col_name == "website":
                cell.value = publisher.website or ""
            elif col_name == "books_count":
                cell.value = books_count
            elif col_name == "created_at":
                cell.value = (
                    publisher.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if publisher.created_at
                    else ""
                )
            elif col_name == "years_active":
                if publisher.founded_year:
                    cell.value = current_year - publisher.founded_year
                else:
                    cell.value = "Unknown"

        row_idx += 1

    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add books sheet if requested
    if include_books:
        books_ws = wb.create_sheet("Books by Publisher")

        # Books sheet headers
        books_headers = [
            "Publisher ID",
            "Publisher Name",
            "Founded Year",
            "Book ID",
            "Book Title",
            "ISBN13",
            "Year Published",
            "Pages",
            "Language",
        ]
        for col_idx, header in enumerate(books_headers, 1):
            cell = books_ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write books data
        books_row = 2
        for publisher in queryset:
            books = publisher.books.all()

            if books:
                for book in books:
                    books_ws.cell(row=books_row, column=1).value = publisher.id
                    books_ws.cell(row=books_row, column=2).value = publisher.name
                    books_ws.cell(row=books_row, column=3).value = (
                        publisher.founded_year or ""
                    )
                    books_ws.cell(row=books_row, column=4).value = book.id
                    books_ws.cell(row=books_row, column=5).value = book.title
                    books_ws.cell(row=books_row, column=6).value = book.isbn13 or ""
                    books_ws.cell(row=books_row, column=7).value = (
                        book.publish_year or ""
                    )
                    books_ws.cell(row=books_row, column=8).value = book.pages or ""
                    books_ws.cell(row=books_row, column=9).value = (
                        book.language_code or ""
                    )
                    books_row += 1
            else:
                # Add publisher row even if no books
                books_ws.cell(row=books_row, column=1).value = publisher.id
                books_ws.cell(row=books_row, column=2).value = publisher.name
                books_ws.cell(row=books_row, column=3).value = (
                    publisher.founded_year or ""
                )
                books_ws.cell(row=books_row, column=4).value = "No books"
                books_row += 1

        # Auto-adjust books sheet column widths
        for col_idx in range(1, len(books_headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in books_ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            books_ws.column_dimensions[column_letter].width = adjusted_width

    return wb
